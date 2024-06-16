import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import pathlib
import openpyxl, xlrd
from openpyxl import Workbook


ctk.set_appearance_mode('dark')
ctk.set_default_color_theme('dark-blue')

#janela
window = ctk.CTk()
window.geometry('500x500')
window.title('Login')

#frame
frame = ctk.CTkFrame(window, width=500, height=100, fg_color='teal')
frame.place(x=0, y=10)

titulo = ctk.CTkLabel(frame, text='Tela de Login', fg_color='transparent', font=('Arial bold', 25), text_color='#fff')
titulo.place(x=180, y=40)

#tema
def mudarTema(novoTema):
    ctk.set_appearance_mode(novoTema)

tema = ctk.CTkOptionMenu(window, values=['Dark', 'System'], command=mudarTema)
tema.place(y=460, x=10)

#corpo
label = ctk.CTkLabel(window, text='Faça seu Login!', fg_color='transparent', text_color=['#000', '#fff'], font=('Arial', 12))
label.place(x=10, y=120)

labelLogin = ctk.CTkLabel(window, text='LOGIN:', text_color=['#000', '#fff'], font=('Arial', 14))
labelLogin.place(x=240, y=140)

entryLogin = ctk.CTkEntry(window, width=250, height=35, placeholder_text='Nome de Usuário', border_color=['teal', '#aaa'])
entryLogin.place(x=140, y=170)

labelSenha = ctk.CTkLabel(window, text='SENHA:', width=250, height=35, text_color=['#000', '#fff'], font=('Arial', 14) )
labelSenha.place(x=140, y=220)

entrySenha = ctk.CTkEntry(window, width=250, height=35, placeholder_text='Senha', border_color=['teal', '#aaa'], show='*')
entrySenha.place(x=140, y=250)

ficheiro = pathlib.Path('Calculadora.xlsx')
if ficheiro.exists():
    pass
else:
    ficheiro = Workbook()
    folha = ficheiro.active
    folha['A1'] = 'Login:'
    folha['B1'] = 'Senha:'
    ficheiro.save('Calculadora.xlsx')

#botões
def enviar():
    nomeUsuario = entryLogin.get()
    senha = entrySenha.get()

    ficheiro = openpyxl.load_workbook('Calculadora.xlsx')
    folha = ficheiro.active
    folha.cell(column=1, row=folha.max_row+1, value=nomeUsuario)
    folha.cell(column=2, row=folha.max_row, value=senha)

    ficheiro.save(r'Calculadora.xlsx')
    messagebox.showinfo('Sistema', 'Dados Salvos')

    def novaJanela():
        calculadora = ctk.CTkToplevel(window)
        calculadora.geometry('700x500')
        calculadora.title('Calculadora')

        #frame
        frameCalculadora = ctk.CTkFrame(calculadora, width=700, height=150, fg_color='#e08402')
        frameCalculadora.place(x=0, y=0)

        tituloCalculadora = ctk.CTkLabel(frameCalculadora, text='Calculadora', fg_color='transparent', font=('Arial bold', 30), text_color='#fff')
        tituloCalculadora.place(x=270, y=60)
        
        #corpo
        labelCalculadora = ctk.CTkLabel(calculadora, text='Digite seu cálculo simples:', text_color=['#000', '#fff'], font=('Arial', 16))
        labelCalculadora.place(x=260, y=170)

        entryCalculadora = ctk.CTkEntry(calculadora, placeholder_text='Cálculo', height=80, width=400, font=('Arial', 20), border_color=['teal', '#aaa'])
        entryCalculadora.place(x=170, y=220)

        #botao
        def calcular():
            tituloCalculadora.configure(text=str(eval(entryCalculadora.get())))

        botaoCalcular = ctk.CTkButton(calculadora, text='Calcular', font=('Arial', 16), height=50, width=300, command=calcular)
        botaoCalcular.place(x=220, y=350)

        def fechar():
            window.destroy()

        botaoFechar = ctk.CTkButton(calculadora, text='Fechar', font=('Arial', 16), height=50, width=300, command=fechar)
        botaoFechar.place(x=220, y=440)

    botaoCalculadora = ctk.CTkButton(window, text='Calculadora', command=novaJanela)
    botaoCalculadora.place(x=190, y=400)

    

botaoLogin = ctk.CTkButton(window, text='Fazer Login', command=enviar)
botaoLogin.place(x=190, y=340)




window.mainloop()
